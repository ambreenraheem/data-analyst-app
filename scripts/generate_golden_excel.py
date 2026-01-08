"""Generate golden dataset Excel file for contract testing."""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

def create_sample_excel():
    """Create sample_excel_report.xlsx with financial data."""
    wb = Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # Create Income Statement sheet
    ws_income = wb.create_sheet("Income Statement")
    create_income_statement(ws_income)

    # Create Balance Sheet sheet
    ws_balance = wb.create_sheet("Balance Sheet")
    create_balance_sheet(ws_balance)

    # Create Cash Flow sheet
    ws_cashflow = wb.create_sheet("Cash Flow")
    create_cash_flow(ws_cashflow)

    # Save file
    output_path = Path(__file__).parent.parent / "tests" / "contract" / "golden_datasets" / "sample_excel_report.xlsx"
    wb.save(output_path)
    print(f"Created: {output_path}")

def create_income_statement(ws):
    """Create Income Statement sheet."""
    # Header
    ws["A1"] = "RetailCo LLC"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = "Income Statement"
    ws["A2"].font = Font(bold=True, size=12)
    ws["A3"] = "For the Year Ended December 31, 2024"
    ws["A3"].font = Font(italic=True)

    # Column headers
    ws["A5"] = "Line Item"
    ws["B5"] = "Amount (USD)"
    ws["A5"].font = Font(bold=True)
    ws["B5"].font = Font(bold=True)
    ws["A5"].fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    ws["B5"].fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    # Data rows (matching expected_outputs.json)
    data = [
        ("Revenue", 42500000),
        ("Cost of Goods Sold", 25500000),
        ("Gross Profit", 17000000),
        ("Operating Expenses", 12000000),
        ("Operating Income", 5000000),
        ("Interest Expense", 300000),
        ("Income Before Tax", 4700000),
        ("Tax Expense", 1300000),
        ("Net Income", 3400000),
    ]

    row = 6
    for label, value in data:
        ws[f"A{row}"] = label
        ws[f"B{row}"] = value
        ws[f"B{row}"].number_format = "$#,##0"
        if label in ["Gross Profit", "Operating Income", "Net Income"]:
            ws[f"A{row}"].font = Font(bold=True)
            ws[f"B{row}"].font = Font(bold=True)
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20

def create_balance_sheet(ws):
    """Create Balance Sheet sheet."""
    # Header
    ws["A1"] = "RetailCo LLC"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = "Balance Sheet"
    ws["A2"].font = Font(bold=True, size=12)
    ws["A3"] = "As of December 31, 2024"
    ws["A3"].font = Font(italic=True)

    # Column headers
    ws["A5"] = "Account"
    ws["B5"] = "Amount (USD)"
    ws["A5"].font = Font(bold=True)
    ws["B5"].font = Font(bold=True)
    ws["A5"].fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    ws["B5"].fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    # Data
    data = [
        ("ASSETS", ""),
        ("Current Assets", 15000000),
        ("Property & Equipment", 8000000),
        ("Total Assets", 23000000),
        ("", ""),
        ("LIABILITIES", ""),
        ("Current Liabilities", 6000000),
        ("Long-term Debt", 5000000),
        ("Total Liabilities", 11000000),
        ("", ""),
        ("EQUITY", ""),
        ("Shareholders' Equity", 12000000),
        ("Total Liabilities & Equity", 23000000),
    ]

    row = 6
    for label, value in data:
        ws[f"A{row}"] = label
        if value != "":
            ws[f"B{row}"] = value
            ws[f"B{row}"].number_format = "$#,##0"
        if label in ["ASSETS", "LIABILITIES", "EQUITY"]:
            ws[f"A{row}"].font = Font(bold=True, underline="single")
        if label in ["Total Assets", "Total Liabilities", "Shareholders' Equity", "Total Liabilities & Equity"]:
            ws[f"A{row}"].font = Font(bold=True)
            ws[f"B{row}"].font = Font(bold=True)
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20

def create_cash_flow(ws):
    """Create Cash Flow sheet."""
    # Header
    ws["A1"] = "RetailCo LLC"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = "Statement of Cash Flows"
    ws["A2"].font = Font(bold=True, size=12)
    ws["A3"] = "For the Year Ended December 31, 2024"
    ws["A3"].font = Font(italic=True)

    # Column headers
    ws["A5"] = "Activity"
    ws["B5"] = "Amount (USD)"
    ws["A5"].font = Font(bold=True)
    ws["B5"].font = Font(bold=True)
    ws["A5"].fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    ws["B5"].fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    # Data
    data = [
        ("OPERATING ACTIVITIES", ""),
        ("Net Income", 3400000),
        ("Depreciation", 800000),
        ("Changes in Working Capital", -500000),
        ("Net Cash from Operating", 3700000),
        ("", ""),
        ("INVESTING ACTIVITIES", ""),
        ("Capital Expenditures", -1200000),
        ("Net Cash from Investing", -1200000),
        ("", ""),
        ("FINANCING ACTIVITIES", ""),
        ("Debt Repayment", -1000000),
        ("Dividends Paid", -800000),
        ("Net Cash from Financing", -1800000),
        ("", ""),
        ("Net Change in Cash", 700000),
    ]

    row = 6
    for label, value in data:
        ws[f"A{row}"] = label
        if value != "":
            ws[f"B{row}"] = value
            ws[f"B{row}"].number_format = "$#,##0"
        if label in ["OPERATING ACTIVITIES", "INVESTING ACTIVITIES", "FINANCING ACTIVITIES"]:
            ws[f"A{row}"].font = Font(bold=True, underline="single")
        if label.startswith("Net Cash") or label == "Net Change in Cash":
            ws[f"A{row}"].font = Font(bold=True)
            ws[f"B{row}"].font = Font(bold=True)
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20

if __name__ == "__main__":
    create_sample_excel()
