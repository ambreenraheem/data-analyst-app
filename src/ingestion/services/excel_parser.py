"""
Excel parser service.

Provides multi-sheet Excel parsing with cell reference tracking
using openpyxl library.

Supports Constitution Principle I (Data-First with source traceability).
"""

import io
import logging
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.utils import get_column_letter

from src.shared.exceptions import ExtractionFailedError


logger = logging.getLogger("excel_parser")
logger.setLevel(logging.INFO)


class ExcelParser:
    """Parser for Excel files with multi-sheet support."""

    @staticmethod
    def extract_tables_from_excel(
        file_bytes: bytes, document_id: str, document_name: str
    ) -> Dict[str, Any]:
        """
        Extract tables from Excel file across all sheets.

        Args:
            file_bytes: Excel file content as bytes
            document_id: Unique document identifier
            document_name: Original filename

        Returns:
            Dictionary containing:
            {
                "tables": [...],  # List of extracted tables (one per sheet)
                "sheet_count": 3,
                "overall_confidence": 1.0  # Excel has perfect confidence
            }

        Raises:
            ExtractionFailedError: If extraction fails
        """
        logger.info(f"Starting Excel extraction for document: {document_id}")

        try:
            # Load workbook from bytes
            file_stream = io.BytesIO(file_bytes)
            workbook = load_workbook(file_stream, data_only=True, read_only=True)

            tables = []

            # Process each sheet
            for sheet_idx, sheet_name in enumerate(workbook.sheetnames):
                sheet = workbook[sheet_name]

                logger.info(f"Processing sheet: {sheet_name} ({sheet_idx + 1}/{len(workbook.sheetnames)})")

                # Extract sheet as a table
                table_data = ExcelParser._extract_sheet_as_table(
                    sheet=sheet,
                    sheet_name=sheet_name,
                    sheet_index=sheet_idx
                )

                if table_data["cells"]:  # Only add non-empty sheets
                    tables.append(table_data)

            workbook.close()

            logger.info(f"Extracted {len(tables)} sheets from Excel document: {document_id}")

            return {
                "tables": tables,
                "sheet_count": len(workbook.sheetnames),
                "overall_confidence": 1.0,  # Excel extraction has perfect confidence
            }

        except Exception as e:
            logger.error(f"Excel extraction failed for document {document_id}: {e}")
            raise ExtractionFailedError(
                document_id=document_id,
                reason=f"Excel parsing failed: {str(e)}",
                retry_eligible=False  # Excel parsing failures are usually not retryable
            )

    @staticmethod
    def _extract_sheet_as_table(
        sheet: Any, sheet_name: str, sheet_index: int
    ) -> Dict[str, Any]:
        """
        Extract a single Excel sheet as a table.

        Args:
            sheet: openpyxl worksheet
            sheet_name: Name of the sheet
            sheet_index: Index of the sheet (0-based)

        Returns:
            Table dictionary with cells and source references
        """
        # Determine used range
        max_row = sheet.max_row
        max_col = sheet.max_column

        # Create table structure
        table_data = {
            "table_id": f"sheet-{sheet_index + 1}",
            "sheet_name": sheet_name,
            "row_count": max_row,
            "column_count": max_col,
            "cells": []
        }

        # Extract all cells
        for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=max_row, max_col=max_col)):
            for col_idx, cell in enumerate(row):
                # Skip empty cells
                if cell.value is None or str(cell.value).strip() == "":
                    continue

                cell_data = ExcelParser._extract_cell_data(
                    cell=cell,
                    row_index=row_idx,
                    column_index=col_idx,
                    sheet_name=sheet_name
                )

                table_data["cells"].append(cell_data)

        return table_data

    @staticmethod
    def _extract_cell_data(
        cell: Cell, row_index: int, column_index: int, sheet_name: str
    ) -> Dict[str, Any]:
        """
        Extract data from a single Excel cell.

        Args:
            cell: openpyxl Cell object
            row_index: Row index (0-based)
            column_index: Column index (0-based)
            sheet_name: Name of the sheet

        Returns:
            Cell dictionary with content and source reference
        """
        # Get cell reference (e.g., "A1", "B34")
        cell_reference = f"{get_column_letter(column_index + 1)}{row_index + 1}"

        # Format content based on cell type
        content = ExcelParser._format_cell_value(cell.value)

        return {
            "row_index": row_index,
            "column_index": column_index,
            "row_span": 1,
            "column_span": 1,
            "content": content,
            "confidence": 1.0,  # Excel has perfect confidence
            "cell_reference": f"{sheet_name}!{cell_reference}",
            "kind": "content"
        }

    @staticmethod
    def _format_cell_value(value: Any) -> str:
        """
        Format cell value as string.

        Args:
            value: Cell value (can be str, int, float, datetime, etc.)

        Returns:
            Formatted string representation
        """
        if value is None:
            return ""

        # Handle different types
        if isinstance(value, (int, float)):
            # Format numbers
            if isinstance(value, float) and value.is_integer():
                return str(int(value))
            return str(value)

        # Convert to string for other types
        return str(value).strip()

    @staticmethod
    def identify_financial_tables(tables: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """
        Identify which sheets likely contain financial data.

        Looks for keywords in sheet names and content.

        Args:
            tables: List of extracted tables

        Returns:
            List of tables likely containing financial data
        """
        financial_keywords = [
            "income", "revenue", "profit", "loss", "balance", "sheet",
            "statement", "financial", "cash", "flow", "p&l", "pl",
            "earnings", "expenses", "assets", "liabilities", "equity"
        ]

        financial_tables = []

        for table in tables:
            sheet_name = table.get("sheet_name", "").lower()

            # Check if sheet name contains financial keywords
            is_financial = any(keyword in sheet_name for keyword in financial_keywords)

            if is_financial:
                financial_tables.append(table)
            else:
                # Check cell content for financial keywords
                content_text = " ".join(
                    [cell.get("content", "").lower() for cell in table.get("cells", [])]
                )
                if any(keyword in content_text for keyword in financial_keywords):
                    financial_tables.append(table)

        logger.info(f"Identified {len(financial_tables)} financial sheets out of {len(tables)} total")

        return financial_tables if financial_tables else tables  # Return all if none identified


# Global parser instance
excel_parser = ExcelParser()
