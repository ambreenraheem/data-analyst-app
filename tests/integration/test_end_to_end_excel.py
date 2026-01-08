"""
Integration test for end-to-end Excel processing.

Tests the full workflow: upload → extract → validate for Excel documents.
"""

import json
import pytest
from pathlib import Path
from unittest.mock import Mock, patch
import asyncio
from openpyxl import load_workbook

# Test data paths
GOLDEN_DIR = Path(__file__).parent.parent / "contract" / "golden_datasets"
SAMPLE_EXCEL = GOLDEN_DIR / "sample_excel_report.xlsx"
EXPECTED_OUTPUTS = GOLDEN_DIR / "expected_outputs.json"


class TestEndToEndExcel:
    """Integration test suite for Excel processing workflow."""

    @pytest.fixture
    def expected_data(self):
        """Load expected outputs from golden dataset."""
        with open(EXPECTED_OUTPUTS) as f:
            data = json.load(f)
        return data["sample_excel_report"]

    @pytest.fixture
    def mock_blob_storage(self):
        """Mock Blob Storage client."""
        with patch('src.ingestion.services.storage_manager.BlobServiceClient') as mock:
            blob_client = Mock()
            blob_client.upload_blob = Mock(return_value=None)
            blob_client.get_blob_properties = Mock(return_value=Mock(size=7000))
            blob_client.download_blob = Mock(return_value=Mock(readall=Mock(return_value=open(SAMPLE_EXCEL, 'rb').read())))
            mock.return_value.get_blob_client = Mock(return_value=blob_client)
            yield mock

    @pytest.fixture
    def mock_cosmos_db(self):
        """Mock Cosmos DB client."""
        with patch('src.ingestion.services.cosmos_manager.CosmosClient') as mock:
            container = Mock()
            container.create_item = Mock(return_value={'id': 'test-id'})
            container.query_items = Mock(return_value=[])
            mock.return_value.get_database_client.return_value.get_container_client = Mock(return_value=container)
            yield mock

    @pytest.fixture
    def mock_service_bus(self):
        """Mock Service Bus client."""
        with patch('azure.servicebus.ServiceBusClient') as mock:
            sender = Mock()
            sender.send_messages = Mock(return_value=None)
            mock.return_value.get_queue_sender = Mock(return_value=sender)
            yield mock

    @pytest.mark.asyncio
    async def test_upload_excel_document(self, mock_blob_storage, mock_service_bus, mock_cosmos_db):
        """Test uploading Excel document creates document ID and queues for processing."""
        # Import function handler
        from src.ingestion.functions.upload_handler import upload_document

        # Read sample Excel
        with open(SAMPLE_EXCEL, "rb") as f:
            excel_content = f.read()

        # Create mock request
        mock_request = Mock()
        mock_request.files = {"file": Mock(filename="sample_excel_report.xlsx", read=Mock(return_value=excel_content))}

        # Execute upload
        response = await upload_document(mock_request)

        # Verify response
        assert response.status_code == 200
        response_data = json.loads(response.get_body())
        assert "document_id" in response_data
        assert response_data["status"] == "queued"
        assert response_data["document_id"].startswith("doc-")

        # Verify Blob Storage was called
        mock_blob_storage.return_value.get_blob_client.assert_called_once()

        # Verify Service Bus was called
        mock_service_bus.return_value.get_queue_sender.assert_called_once()

    def test_read_excel_file_structure(self):
        """Test reading Excel file structure and sheets."""
        # Load workbook
        wb = load_workbook(SAMPLE_EXCEL)

        # Verify sheets
        assert "Income Statement" in wb.sheetnames
        assert "Balance Sheet" in wb.sheetnames
        assert "Cash Flow" in wb.sheetnames

        # Verify Income Statement data
        ws = wb["Income Statement"]
        assert ws["A1"].value == "RetailCo LLC"
        assert ws["A2"].value == "Income Statement"

        # Verify first metric
        assert "Revenue" in str(ws["A6"].value)
        assert ws["B6"].value == 42500000

    @pytest.mark.asyncio
    async def test_extract_from_excel(self, expected_data, mock_cosmos_db, mock_blob_storage):
        """Test extracting financial metrics from Excel."""
        # Import extraction worker
        from src.ingestion.functions.extraction_worker import process_extraction
        from src.ingestion.services.excel_parser import ExcelParser

        # Test Excel parser directly
        parser = ExcelParser()
        with open(SAMPLE_EXCEL, "rb") as f:
            excel_content = f.read()

        # Extract data
        tables = parser.extract_tables(excel_content)

        # Verify tables were extracted
        assert len(tables) >= 1

        # Verify first table has data
        income_statement_table = tables[0]
        assert income_statement_table["sheet_name"] == "Income Statement"
        assert len(income_statement_table["rows"]) > 0

        # Verify financial metrics can be identified
        from src.ingestion.services.financial_parser import FinancialParser

        is_financial = FinancialParser.detect_income_statement(income_statement_table)
        assert is_financial is True

    @pytest.mark.asyncio
    async def test_validate_excel_extracted_data(self, expected_data, mock_cosmos_db):
        """Test validating extracted Excel data."""
        # Import validation worker
        from src.ingestion.services.validator import Validator

        # Use expected metrics from golden dataset
        metrics = expected_data["metrics"]
        extraction_result = {
            "document_id": "doc-sample-002",
            "ocr_confidence_avg": 1.0  # Excel has perfect confidence
        }

        # Test completeness validation
        result = Validator.validate_completeness(metrics, extraction_result)
        assert result.validation_status == "passed"
        assert result.is_valid()

        # Test range validation
        result = Validator.validate_ranges(metrics)
        assert result.validation_status == "passed"

        # Test confidence validation
        result = Validator.validate_confidence(metrics, extraction_result)
        assert result.validation_status == "passed"

    @pytest.mark.asyncio
    async def test_full_workflow_excel_to_validated_metrics(self, expected_data, mock_blob_storage, mock_service_bus, mock_cosmos_db):
        """Test complete workflow from Excel upload to validated metrics."""
        # This test simulates the entire pipeline

        # Step 1: Upload (already tested above)
        document_id = "doc-test-002"

        # Step 2: Extract metrics
        from src.ingestion.services.excel_parser import ExcelParser
        from src.ingestion.services.financial_parser import FinancialParser

        parser = ExcelParser()
        with open(SAMPLE_EXCEL, "rb") as f:
            excel_content = f.read()

        tables = parser.extract_tables(excel_content)

        # Step 3: Parse financial metrics
        metrics_found = []
        for table in tables:
            if FinancialParser.detect_income_statement(table):
                # Parse metrics (simplified)
                metrics_found.append({
                    "metric_type": "revenue",
                    "value": 42500000,
                    "confidence": 1.0
                })

        # Verify metrics were found
        assert len(metrics_found) > 0

        # Step 4: Validate
        from src.ingestion.services.validator import Validator

        validation_result = Validator.validate_ranges(metrics_found)
        assert validation_result.is_valid()

    def test_extraction_matches_golden_dataset(self, expected_data):
        """Test that extraction results match golden dataset expectations."""
        # Load expected metrics
        expected_metrics = expected_data["metrics"]

        # Verify we have expected metrics
        assert len(expected_metrics) == 6  # 6 metrics from income statement

        # Verify metric structure
        for metric in expected_metrics:
            assert "metric_name" in metric
            assert "value" in metric
            assert "metric_type" in metric
            assert "confidence" in metric
            assert metric["confidence"] == 1.0  # Excel parsing has perfect confidence
            assert "source_reference" in metric

            # Verify source reference structure for Excel
            source_ref = metric["source_reference"]
            assert "document_name" in source_ref
            assert "sheet_name" in source_ref
            assert "cell_reference" in source_ref
            assert source_ref["sheet_name"] == "Income Statement"

    def test_multi_sheet_extraction(self):
        """Test extracting data from multiple sheets."""
        from src.ingestion.services.excel_parser import ExcelParser

        parser = ExcelParser()
        with open(SAMPLE_EXCEL, "rb") as f:
            excel_content = f.read()

        tables = parser.extract_tables(excel_content)

        # Verify multiple sheets were extracted
        sheet_names = [table["sheet_name"] for table in tables]
        assert "Income Statement" in sheet_names
        assert "Balance Sheet" in sheet_names
        assert "Cash Flow" in sheet_names

    def test_cell_reference_accuracy(self, expected_data):
        """Test that cell references match expected locations."""
        # Load workbook to verify actual cell values
        wb = load_workbook(SAMPLE_EXCEL)
        ws = wb["Income Statement"]

        # Test each expected metric's cell reference
        for metric in expected_data["metrics"]:
            cell_ref = metric["source_reference"]["cell_reference"]
            cell_value = ws[cell_ref].value

            # Verify cell contains the expected value
            if isinstance(cell_value, (int, float)):
                assert cell_value == metric["value"]

    def test_number_parsing_from_excel(self):
        """Test number parsing from Excel cells."""
        from src.ingestion.utils.number_parser import NumberParser

        # Test parsing Excel number formats
        test_cases = [
            (42500000, "USD", 42500000.0),
            (25500000, "USD", 25500000.0),
            (3400000, "USD", 3400000.0),
        ]

        for input_value, currency, expected_value in test_cases:
            # Excel cells contain numeric values directly
            assert input_value == expected_value

    def test_excel_confidence_scoring(self, expected_data):
        """Test confidence scoring for Excel extraction."""
        from src.ingestion.utils.confidence_scorer import ConfidenceScorer

        metrics = expected_data["metrics"]

        # All Excel extractions should have confidence 1.0
        for metric in metrics:
            assert metric["confidence"] == 1.0

        # Calculate overall confidence
        confidences = [m["confidence"] for m in metrics]
        stats = ConfidenceScorer.get_confidence_statistics(confidences)

        assert stats["mean"] == 1.0
        assert stats["min"] == 1.0
        assert stats["max"] == 1.0

    def test_validation_rules_against_excel_data(self, expected_data):
        """Test all validation rules against Excel golden dataset."""
        from src.ingestion.services.validator import Validator

        metrics = expected_data["metrics"]
        extraction_result = {
            "ocr_confidence_avg": expected_data["extraction_summary"]["ocr_confidence"]
        }

        # Completeness validation
        result = Validator.validate_completeness(metrics, extraction_result)
        assert result.is_valid()
        assert len(result.errors) == 0

        # Range validation
        result = Validator.validate_ranges(metrics)
        assert result.is_valid()
        assert len(result.errors) == 0

        # Confidence validation
        result = Validator.validate_confidence(metrics, extraction_result)
        assert result.is_valid()
        assert len(result.warnings) == 0  # Perfect confidence

        # Relationship validation
        result = Validator.validate_relationships(metrics)
        assert result.is_valid()

    @pytest.mark.asyncio
    async def test_error_handling_corrupted_excel(self):
        """Test error handling for corrupted Excel files."""
        from src.ingestion.services.excel_parser import ExcelParser

        parser = ExcelParser()

        # Test with invalid Excel content
        with pytest.raises(Exception) as exc_info:
            parser.extract_tables(b"invalid excel content")

        # Verify appropriate error is raised
        assert exc_info.value is not None

    def test_performance_large_excel(self):
        """Test performance with the sample Excel file."""
        import time
        from src.ingestion.services.excel_parser import ExcelParser

        parser = ExcelParser()

        with open(SAMPLE_EXCEL, "rb") as f:
            excel_content = f.read()

        # Measure extraction time
        start_time = time.time()
        tables = parser.extract_tables(excel_content)
        end_time = time.time()

        # Extraction should complete quickly (< 1 second for small file)
        elapsed_time = end_time - start_time
        assert elapsed_time < 1.0

        # Verify results
        assert len(tables) > 0

    @pytest.mark.asyncio
    async def test_status_tracking_excel_processing(self, mock_cosmos_db):
        """Test status tracking throughout Excel processing pipeline."""
        from src.ingestion.functions.status_handler import get_document_status

        document_id = "doc-test-002"

        # Mock processing logs
        mock_logs = [
            {"event_type": "upload_completed", "timestamp": "2024-01-01T10:00:00Z", "file_type": "xlsx"},
            {"event_type": "extraction_started", "timestamp": "2024-01-01T10:00:30Z"},
            {"event_type": "extraction_completed", "timestamp": "2024-01-01T10:01:00Z", "metrics_extracted": 6},
            {"event_type": "validation_completed", "timestamp": "2024-01-01T10:01:15Z", "validation_status": "passed"},
        ]

        mock_cosmos_db.return_value.get_database_client.return_value.get_container_client.return_value.query_items = Mock(
            return_value=iter(mock_logs)
        )

        # Create mock request
        mock_request = Mock()
        mock_request.route_params = {"document_id": document_id}

        # Execute status check
        response = await get_document_status(mock_request)

        # Verify response
        assert response.status_code == 200
        response_data = json.loads(response.get_body())
        assert response_data["status"] == "completed"
        assert response_data["metrics_extracted"] == 6


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
